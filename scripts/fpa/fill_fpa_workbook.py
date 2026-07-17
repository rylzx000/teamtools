#!/usr/bin/env python
"""Fill the official FPA workbook template from a structured payload JSON."""

from __future__ import annotations

import argparse
import copy
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE_PATH = ROOT_DIR / "data/modules/fpa/profile/templates/fpa_template.xlsx"
DEFAULT_MAPPING_PATH = ROOT_DIR / "data/modules/fpa/profile/mapping/excel_mapping.yaml"

SHEET_PROJECT = "项目特征"
SHEET_SIZE = "规模估算"
SHEET_COST = "开发费用估算"
SHEET_PARAMS = "模板使用说明&基础参数"

DETAIL_START_ROW = 6
TEMPLATE_DETAIL_END_ROW = 14
TEMPLATE_SUMMARY_ROW = 16
MIN_DETAIL_ROWS = 1

CATEGORIES = {"ILF", "EIF", "EI", "EO", "EQ"}
REUSE_FACTORS = {"高": 1 / 3, "中": 2 / 3, "低": 1.0}
CHANGE_FACTORS = {"新增": 1.0, "修改": 0.8, "删除": 0.2}
ESTIMATION_WEIGHTS = {"ILF": 10.0, "EIF": 7.0, "EI": 4.0, "EO": 5.0, "EQ": 4.0}
PRESIZE_WEIGHTS = {"ILF": 35.0, "EIF": 15.0, "EI": 0.0, "EO": 0.0, "EQ": 0.0}

DEFAULT_PROJECT_FEATURES = {
    "规模计数时机": "估算中期",
    "完整性级别": "完整性级别为A/B同时为达成完整性级别要求采取了特殊的设计及实现方式",
}

COUNT_TIMING_FACTORS = {
    "估算早期": 1.39,
    "估算中期": 1.21,
    "估算晚期": 1.10,
    "项目交付后及运维阶段": 1.00,
}
APPLICATION_TYPE_FACTORS = {
    "业务处理": 1.00,
    "信息管理": 1.00,
    "实时控制": 1.20,
    "嵌入式": 1.30,
    "科学计算": 1.10,
    "系统软件": 1.20,
    "其他": 1.00,
}
QUALITY_RATING_SCORES = {
    "无影响": 0,
    "很低": 1,
    "低": 2,
    "中等": 3,
    "高": 4,
    "很高": 5,
    "极高": 5,
}
INTEGRITY_FACTORS = {
    "没有明确的完整性级别或等级为C/D": 1.00,
    "完整性级别为A/B同时为达成完整性级别要求采取了特殊的设计及实现方式": 1.10,
    "完整性级别为A同时为达成完整性级别要求在软件开发全生命周期均采取了特定、明确的措施": 1.30,
}
LANGUAGE_FACTORS = {
    "4GL": 0.80,
    "3GL": 1.00,
    "2GL": 1.20,
}
TEAM_FACTORS = {
    "熟悉": 0.90,
    "中等": 1.00,
    "陌生": 1.10,
}
PLATFORM_FACTORS = {
    "通用平台": 1.00,
    "专用平台": 1.10,
}
PRODUCTIVITY = {"low": 6.32, "middle": 7.88, "high": 9.45}

PROJECT_FEATURE_CELL_MAP = {
    "规模计数时机": "C1",
    "应用类型": "C2",
    "分布式处理": "C3",
    "性能": "C4",
    "可靠性": "C5",
    "多重站点": "C6",
    "完整性级别": "C7",
    "开发语言": "C8",
    "开发团队背景": "C9",
    "开发平台": "C10",
}

PARAM_FACTOR_RANGES = {
    "规模计数时机": ("H50", "I53"),
    "应用类型": ("G10", "I16"),
    "分布式处理": ("H19", "I21"),
    "性能": ("H22", "I24"),
    "可靠性": ("H25", "I27"),
    "多重站点": ("H28", "I30"),
    "完整性级别": ("H45", "I47"),
    "开发语言": ("H34", "I36"),
    "开发团队背景": ("H40", "I42"),
    "开发平台": ("H56", "I57"),
}

PROJECT_FEATURE_FACTOR_FALLBACKS = {
    "规模计数时机": COUNT_TIMING_FACTORS,
    "完整性级别": INTEGRITY_FACTORS,
}

ITEM_COLUMN_MAP = {
    "system": "B",
    "level1_module": "C",
    "level2_module": "D",
    "level3_module": "E",
    "level4_module": "F",
    "function_description": "G",
    "count_item_name": "H",
    "category": "I",
    "reuse": "K",
    "change_type": "L",
    "remark": "N",
}

FEATURE_ALIASES = {
    "count_timing": "规模计数时机",
    "application_type": "应用类型",
    "distributed_processing": "分布式处理",
    "performance": "性能",
    "reliability": "可靠性",
    "multi_site": "多重站点",
    "integrity_level": "完整性级别",
    "development_language": "开发语言",
    "team_background": "开发团队背景",
    "development_platform": "开发平台",
}


class FpaPayloadError(ValueError):
    """Raised when the payload cannot generate a valid workbook."""


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise FpaPayloadError("payload JSON 顶层必须是对象")
    return data


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def resolve_path(value: str | None, base_dir: Path) -> Path | None:
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = base_dir / path
    return path.resolve()


def resolve_mapping_path(value: str | None) -> Path:
    if not value:
        return DEFAULT_TEMPLATE_PATH
    path = Path(value)
    if not path.is_absolute():
        path = ROOT_DIR / path
    return path.resolve()


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip()
    return cleaned or "FPA工作量评估"


def load_mapping(path: Path | None) -> dict[str, Any]:
    """Load a tiny YAML subset used for human-readable mapping metadata."""

    if not path or not path.exists():
        return {}
    mapping: dict[str, Any] = {}
    stack: list[tuple[int, dict[str, Any]]] = [(-1, mapping)]
    for raw in path.read_text(encoding="utf-8").splitlines():
        if not raw.strip() or raw.lstrip().startswith("#"):
            continue
        indent = len(raw) - len(raw.lstrip(" "))
        line = raw.strip()
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip().strip('"').strip("'")
        value = value.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()
        parent = stack[-1][1]
        if value == "":
            child: dict[str, Any] = {}
            parent[key] = child
            stack.append((indent, child))
        else:
            parent[key] = value.strip('"').strip("'")
    return mapping


def normalize_project_features(payload: dict[str, Any]) -> dict[str, str]:
    project_features = dict(DEFAULT_PROJECT_FEATURES)
    overrides = payload.get("project_features") or {}
    if not isinstance(overrides, dict):
        raise FpaPayloadError("project_features 必须是对象")
    for key, value in overrides.items():
        normalized_key = FEATURE_ALIASES.get(str(key), str(key))
        if normalized_key in DEFAULT_PROJECT_FEATURES and value not in (None, ""):
            project_features[normalized_key] = str(value)
    return project_features


def normalize_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise FpaPayloadError("items 必须是非空数组")
    normalized: list[dict[str, str]] = []
    for index, raw in enumerate(items, 1):
        if not isinstance(raw, dict):
            raise FpaPayloadError(f"items[{index}] 必须是对象")
        item: dict[str, Any] = {}
        for key in ITEM_COLUMN_MAP:
            value = raw.get(key, "")
            item[key] = "" if value is None else str(value).strip()
        for trace_key in ("stable_id", "fact_ids", "route_ids", "system_scene_ids"):
            if trace_key in raw:
                item[trace_key] = raw[trace_key]
        for required in ["system", "level1_module", "function_description", "count_item_name"]:
            if not item[required]:
                raise FpaPayloadError(f"items[{index}].{required} 不能为空")
        if item["category"] not in CATEGORIES:
            raise FpaPayloadError(f"items[{index}].category 必须是 ILF/EIF/EI/EO/EQ")
        if item["reuse"] not in REUSE_FACTORS:
            raise FpaPayloadError(f"items[{index}].reuse 必须是 高/中/低")
        if item["change_type"] not in CHANGE_FACTORS:
            raise FpaPayloadError(f"items[{index}].change_type 必须是 新增/修改/删除")
        item["remark"] = build_structured_remark(item)
        normalized.append(item)
    return normalized


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def build_structured_remark(item: dict[str, str]) -> str:
    remark = normalize_space(item.get("remark", ""))
    category = item["category"]
    reuse = item["reuse"]
    change_type = item["change_type"]

    if all(label in remark for label in ["类别原因：", "复用原因：", "修改类型原因："]):
        return normalize_space(remark)

    description = item.get("function_description") or item.get("count_item_name") or "该功能"
    count_item_name = item.get("count_item_name") or "该计数项"
    evidence = f"{description}（{count_item_name}）"
    if remark:
        evidence = f"{evidence}，原始说明：{remark}"

    category_reason_templates = {
        "ILF": f"{evidence}，用于维护系统内部持久业务数据或状态记录，属于内部逻辑文件",
        "EIF": f"{evidence}，依赖外部系统维护的数据或接口资料作为参考，属于外部接口文件",
        "EI": f"{evidence}，由用户或外部系统输入并触发业务数据维护或状态变更，属于外部输入",
        "EO": f"{evidence}，包含派生处理、通知、导出、统计或风险提示等输出，属于外部输出",
        "EQ": f"{evidence}，按条件读取并展示已有数据，未产生复杂派生处理，属于外部查询",
    }
    reuse_reason_templates = {
        "高": f"主要复用既有流程、页面、接口或数据结构，仅需少量适配",
        "中": f"可复用部分既有能力，但需要新增或调整字段、规则、关联关系或处理逻辑",
        "低": f"本次能力相对独立，需要新增主要业务流程、数据结构或处理逻辑",
    }
    change_reason_templates = {
        "新增": f"本次新增 {count_item_name} 相关能力",
        "修改": f"本次在既有 {count_item_name} 能力基础上调整业务规则、字段或处理流程",
        "删除": f"本次移除或停用 {count_item_name} 相关能力",
    }

    return (
        f"类别原因：{category}，{category_reason_templates[category]}；"
        f"复用原因：{reuse}，{reuse_reason_templates[reuse]}；"
        f"修改类型原因：{change_type}，{change_reason_templates[change_type]}。"
    )


def copy_row_style_and_formulas(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)
        if source.value is not None:
            if isinstance(source.value, ArrayFormula):
                target.value = None
            elif isinstance(source.value, str) and source.value.startswith("="):
                origin = f"{get_column_letter(column)}{source_row}"
                target.value = Translator(source.value, origin=origin).translate_formula(
                    f"{get_column_letter(column)}{target_row}"
                )
            elif column in (1, 10, 13):
                target.value = source.value


def detail_template_row(row: int) -> int:
    return row if row <= TEMPLATE_DETAIL_END_ROW else TEMPLATE_DETAIL_END_ROW


def preserve_size_sheet_dimensions(ws, row: int) -> None:
    ws.cell(row, 15).value = ""


def clear_detail_inputs(ws, row: int) -> None:
    for column in list(range(2, 10)) + [11, 12, 14]:
        ws.cell(row, column).value = None


def remove_merged_ranges_on_rows(ws, rows: set[int]) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row in rows or merged.max_row in rows:
            ws.unmerge_cells(str(merged))


def remove_merged_ranges_intersecting_rows(ws, start_row: int, end_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= end_row and merged.max_row >= start_row:
            ws.unmerge_cells(str(merged))


def copy_summary_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)


def ensure_detail_rows(ws, item_count: int) -> tuple[int, int]:
    detail_count = max(item_count, MIN_DETAIL_ROWS)
    target_last_detail_row = DETAIL_START_ROW + detail_count - 1
    summary_style_source_row = TEMPLATE_SUMMARY_ROW
    summary_style_snapshot = [copy.copy(ws.cell(summary_style_source_row, column)._style) for column in range(1, ws.max_column + 1)]

    if target_last_detail_row >= TEMPLATE_SUMMARY_ROW:
        remove_merged_ranges_on_rows(ws, {TEMPLATE_SUMMARY_ROW})
        rows_to_insert = target_last_detail_row - TEMPLATE_SUMMARY_ROW + 1
        ws.insert_rows(TEMPLATE_SUMMARY_ROW, rows_to_insert)
        for row in range(TEMPLATE_SUMMARY_ROW, TEMPLATE_SUMMARY_ROW + rows_to_insert):
            copy_row_style_and_formulas(ws, TEMPLATE_DETAIL_END_ROW, row)

    summary_row = target_last_detail_row + 1
    if summary_row < TEMPLATE_SUMMARY_ROW:
        ws.delete_rows(summary_row, TEMPLATE_SUMMARY_ROW - summary_row)

    remove_merged_ranges_intersecting_rows(ws, DETAIL_START_ROW, target_last_detail_row)
    remove_merged_ranges_on_rows(ws, {summary_row})
    copy_summary_style(ws, TEMPLATE_SUMMARY_ROW if summary_row == TEMPLATE_SUMMARY_ROW else summary_row, summary_row)
    for column, style in enumerate(summary_style_snapshot, 1):
        ws.cell(summary_row, column)._style = copy.copy(style)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=9)
    ws.merge_cells(start_row=summary_row, start_column=11, end_row=summary_row, end_column=12)

    for row in range(DETAIL_START_ROW, target_last_detail_row + 1):
        copy_row_style_and_formulas(ws, detail_template_row(row), row)
        clear_detail_inputs(ws, row)

    ws.cell(summary_row, 1).value = "合计"
    ws.cell(summary_row, 10).value = f"=SUM(J{DETAIL_START_ROW}:J{target_last_detail_row})"
    ws.cell(summary_row, 13).value = f"=SUM(M{DETAIL_START_ROW}:M{target_last_detail_row})"
    ws.cell(2, 3).value = f"=J{summary_row}"
    ws.cell(3, 3).value = f"=M{summary_row}"
    preserve_size_sheet_dimensions(ws, summary_row)

    for dv in ws.data_validations.dataValidation:
        formula = str(dv.formula1)
        if "ILF,EIF,EI,EO,EQ" in formula:
            dv.sqref = f"I{DETAIL_START_ROW}:I{target_last_detail_row}"
        elif "高,中,低" in formula:
            dv.sqref = f"K{DETAIL_START_ROW}:K{target_last_detail_row}"
        elif "新增,修改,删除" in formula:
            dv.sqref = f"L{DETAIL_START_ROW}:L{target_last_detail_row}"

    return target_last_detail_row, summary_row


def fill_project_features(ws, features: dict[str, str]) -> None:
    for feature_name, value in features.items():
        cell_address = PROJECT_FEATURE_CELL_MAP.get(feature_name)
        if cell_address:
            ws[cell_address] = value


def fill_items(ws, items: list[dict[str, str]]) -> None:
    for offset, item in enumerate(items):
        row = DETAIL_START_ROW + offset
        for key, column_letter in ITEM_COLUMN_MAP.items():
            ws[f"{column_letter}{row}"] = item[key]


def calculate_item_values(items: list[dict[str, str]], count_mode: str) -> list[dict[str, Any]]:
    weights = PRESIZE_WEIGHTS if count_mode == "预估功能点" else ESTIMATION_WEIGHTS
    calculated: list[dict[str, Any]] = []
    for item in items:
        category = item["category"]
        ufp = weights.get(category, 0.0)
        us = ufp * REUSE_FACTORS[item["reuse"]] * CHANGE_FACTORS[item["change_type"]]
        calculated.append({**item, "ufp": round(ufp, 4), "us": round(us, 4)})
    return calculated


def feature_factor(features: dict[str, str], key: str, table: dict[str, float], default: float = 1.0) -> float:
    return float(table.get(features.get(key, ""), default))


def quality_score(features: dict[str, str], key: str) -> int:
    return int(QUALITY_RATING_SCORES.get(features.get(key, ""), 3))


def read_factor_from_params(ws_params, feature_name: str, selected_value: str, default: float = 1.0) -> float:
    start_cell, end_cell = PARAM_FACTOR_RANGES[feature_name]
    min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
    for row in range(min_row, max_row + 1):
        label = ws_params.cell(row, min_col).value
        factor = ws_params.cell(row, max_col).value
        if label == selected_value:
            try:
                return float(factor)
            except (TypeError, ValueError):
                return default
    return default


def read_effective_project_features(ws_project, ws_params, overrides: dict[str, str]) -> tuple[dict[str, str], dict[str, float]]:
    features: dict[str, str] = {}
    factors: dict[str, float] = {}
    for feature_name, cell_address in PROJECT_FEATURE_CELL_MAP.items():
        value = ws_project[cell_address].value
        if isinstance(value, ArrayFormula):
            value = None
        if overrides.get(feature_name):
            value = overrides[feature_name]
        features[feature_name] = "" if value is None else str(value)
        fallback = PROJECT_FEATURE_FACTOR_FALLBACKS.get(feature_name, {}).get(features[feature_name], 1.0)
        factors[feature_name] = read_factor_from_params(ws_params, feature_name, features[feature_name], fallback)
    return features, factors


def calculate_estimates(
    calculated_items: list[dict[str, Any]],
    features: dict[str, str],
    feature_factors: dict[str, float],
    target_work_days: float | None,
) -> dict[str, Any]:
    adjusted_fp_total = sum(float(item["us"]) for item in calculated_items)
    count_timing_factor = feature_factors.get("规模计数时机", 1.21)
    adjusted_size = adjusted_fp_total * count_timing_factor
    quality_factor = 1 + 0.025 * sum(
        feature_factors.get(key, 0.0)
        for key in ["分布式处理", "性能", "可靠性", "多重站点"]
    )
    application_factor = feature_factors.get("应用类型", 1.0)
    integrity_factor = feature_factors.get("完整性级别", 1.0)
    language_factor = feature_factors.get("开发语言", 1.0)
    team_factor = feature_factors.get("开发团队背景", 1.0)
    platform_factor = feature_factors.get("开发平台", 1.0)

    def work_days(productivity: float) -> float:
        unadjusted = adjusted_size * productivity / 8
        return (
            unadjusted
            * application_factor
            * quality_factor
            * integrity_factor
            * language_factor
            * team_factor
            * platform_factor
        )

    low = work_days(PRODUCTIVITY["low"])
    middle = work_days(PRODUCTIVITY["middle"])
    high = work_days(PRODUCTIVITY["high"])

    target_check = {
        "has_target": target_work_days is not None,
        "target_work_days": target_work_days,
        "hit_status": "not_provided",
        "difference_days": None,
        "difference_ratio": None,
    }
    if target_work_days is not None:
        diff = middle - target_work_days
        ratio = diff / target_work_days if target_work_days else None
        target_check.update(
            {
                "hit_status": "hit" if low <= target_work_days <= high else "out_of_range",
                "difference_days": round(diff, 2),
                "difference_ratio": round(ratio, 4) if ratio is not None else None,
            }
        )

    return {
        "function_point_total": round(sum(float(item["ufp"]) for item in calculated_items), 4),
        "adjusted_fp_total": round(adjusted_fp_total, 4),
        "adjusted_size": round(adjusted_size, 4),
        "productivity": PRODUCTIVITY,
        "factors": {
            "count_timing_factor": round(count_timing_factor, 4),
            "application_type_factor": round(application_factor, 4),
            "quality_factor": round(quality_factor, 4),
            "integrity_level_factor": round(integrity_factor, 4),
            "development_language_factor": round(language_factor, 4),
            "team_background_factor": round(team_factor, 4),
            "development_platform_factor": round(platform_factor, 4),
        },
        "work_days": {
            "low": round(low, 2),
            "middle": round(middle, 2),
            "high": round(high, 2),
        },
        "target_check": target_check,
    }


def build_quality_warnings(
    items: list[dict[str, Any]],
    estimates: dict[str, Any],
    target_work_days: float | None,
) -> tuple[list[dict[str, str]], dict[str, Any], bool]:
    warnings: list[dict[str, str]] = []
    categories = Counter(item["category"] for item in items)
    item_count = len(items)

    if item_count < 3:
        warnings.append(
            {
                "code": "ITEM_COUNT_TOO_LOW",
                "level": "high",
                "message": "功能点条目数过少，存在过度合并或漏拆风险。",
                "suggestion": "复核数据组、输入、输出、查询和接口类基本过程是否完整拆分。",
            }
        )
    if categories.get("ILF", 0) == 0:
        warnings.append(
            {
                "code": "NO_ILF",
                "level": "medium",
                "message": "明细中缺少 ILF，可能遗漏业务数据组或状态记录。",
                "suggestion": "复核是否存在需要维护或长期保存的数据组。",
            }
        )
    if categories.get("EO", 0) == 0:
        warnings.append(
            {
                "code": "NO_EO",
                "level": "medium",
                "message": "明细中缺少 EO，可能遗漏派生输出、通知、报表或风险提示。",
                "suggestion": "复核是否存在计算后输出、导出、推送或提示类功能。",
            }
        )
    if target_work_days is not None and estimates["target_check"]["hit_status"] == "out_of_range":
        warnings.append(
            {
                "code": "TARGET_OUT_OF_RANGE",
                "level": "medium",
                "message": "目标人天不在脚本估算的低-高区间内。",
                "suggestion": "复核目标人天、功能点拆分和项目特征参数是否需要调整。",
            }
        )

    failed_codes = [warning["code"] for warning in warnings if warning["level"] == "high"]
    if failed_codes:
        quality_gate = {
            "status": "failed",
            "failed": True,
            "deliverable_valid": False,
            "reason_codes": failed_codes,
            "required_action": "不建议不复核直接交付；请复核功能点拆分后重新生成或人工确认。",
        }
        deliverable_valid = False
    elif warnings:
        quality_gate = {
            "status": "review_required",
            "failed": False,
            "deliverable_valid": True,
            "reason_codes": [warning["code"] for warning in warnings],
            "required_action": "建议人工复核质量提示后交付。",
        }
        deliverable_valid = True
    else:
        quality_gate = {
            "status": "passed",
            "failed": False,
            "deliverable_valid": True,
            "reason_codes": [],
            "required_action": "无需额外质量门禁动作。",
        }
        deliverable_valid = True

    return warnings, quality_gate, deliverable_valid


def coerce_target_days(payload: dict[str, Any]) -> float | None:
    raw = payload.get("target_work_days", payload.get("target_person_days"))
    if raw in (None, ""):
        return None
    try:
        return float(raw)
    except (TypeError, ValueError) as exc:
        raise FpaPayloadError("target_work_days 必须是数字") from exc


def fill_workbook(
    payload_path: Path,
    output_path_override: Path | None,
    template_path_override: Path | None,
    mapping_path: Path | None,
    process_output_override: Path | None,
) -> dict[str, Any]:
    payload = read_json(payload_path)
    base_dir = payload_path.parent
    requirement_name = str(payload.get("requirement_name") or "").strip()
    if not requirement_name:
        raise FpaPayloadError("requirement_name 是脚本 payload 必填字段")

    mapping = load_mapping(mapping_path or DEFAULT_MAPPING_PATH)
    template_path = (
        template_path_override
        or resolve_path(payload.get("template_path"), base_dir)
        or resolve_mapping_path(mapping.get("workbook", {}).get("default_template"))
    )
    output_path = (
        output_path_override
        or resolve_path(payload.get("output_path"), base_dir)
        or (base_dir / f"{safe_filename(requirement_name)}-FPA工作量评估.xlsx")
    )
    process_output_path = (
        process_output_override
        or resolve_path(payload.get("process_output_path"), base_dir)
        or output_path.with_name(f"{safe_filename(requirement_name)}-FPA生成过程.json")
    )

    if not template_path.exists():
        raise FpaPayloadError(f"模板不存在: {template_path}")

    project_features = normalize_project_features(payload)
    items = normalize_items(payload)
    target_work_days = coerce_target_days(payload)
    count_mode = str(payload.get("count_mode") or payload.get("计数类型") or "估算功能点")
    if count_mode not in {"估算功能点", "预估功能点"}:
        raise FpaPayloadError("count_mode 必须是 估算功能点 或 预估功能点")

    wb = load_workbook(template_path, data_only=False)
    required_sheets = [SHEET_PROJECT, SHEET_SIZE, SHEET_COST, SHEET_PARAMS]
    missing = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
    if missing:
        raise FpaPayloadError(f"模板缺少必要 sheet: {', '.join(missing)}")

    ws_project = wb[SHEET_PROJECT]
    ws_size = wb[SHEET_SIZE]
    ws_cost = wb[SHEET_COST]
    ws_params = wb[SHEET_PARAMS]

    last_detail_row, summary_row = ensure_detail_rows(ws_size, len(items))
    fill_project_features(ws_project, project_features)
    effective_project_features, project_feature_factors = read_effective_project_features(
        ws_project, ws_params, project_features
    )
    ws_size["C1"] = count_mode
    fill_items(ws_size, items)

    ws_cost["C1"] = f"='{SHEET_SIZE}'!M{summary_row}"
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    calculated_items = calculate_item_values(items, count_mode)
    estimates = calculate_estimates(
        calculated_items,
        effective_project_features,
        project_feature_factors,
        target_work_days,
    )
    quality_warnings, quality_gate, deliverable_valid = build_quality_warnings(
        calculated_items, estimates, target_work_days
    )
    process = {
        "schema_version": "fpa.excel_process.v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "requirement_name": requirement_name,
        "assessor": payload.get("assessor"),
        "assessment_date": payload.get("assessment_date"),
        "template_path": str(template_path),
        "output_path": str(output_path),
        "payload_path": str(payload_path),
        "process_output_path": str(process_output_path),
        "count_mode": count_mode,
        "target_work_days": target_work_days,
        "item_count": len(items),
        "detail_range": f"{SHEET_SIZE}!B{DETAIL_START_ROW}:N{last_detail_row}",
        "summary_row": summary_row,
        "project_features": effective_project_features,
        "project_feature_factors": project_feature_factors,
        "items": calculated_items,
        "category_summary": dict(Counter(item["category"] for item in calculated_items)),
        "estimates": estimates,
        "quality_warnings": quality_warnings,
        "quality_gate": quality_gate,
        "deliverable_valid": deliverable_valid,
        "assessment_context": payload.get("assessment_context") or {},
    }
    if payload.get("keep_intermediate_artifacts", True) is not False:
        write_json(process_output_path, process)
    return process


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill FPA workbook template from payload JSON.")
    parser.add_argument("--payload", required=True, help="Excel 脚本输入 payload JSON 路径")
    parser.add_argument("--template", help="可选模板路径；默认使用 profile/templates/fpa_template.xlsx")
    parser.add_argument("--output", help="可选输出 Excel 路径")
    parser.add_argument("--mapping", default=str(DEFAULT_MAPPING_PATH), help="可选 Excel 映射 YAML 路径")
    parser.add_argument("--process-output", help="可选 FPA生成过程.json 输出路径")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        process = fill_workbook(
            payload_path=Path(args.payload).resolve(),
            output_path_override=Path(args.output).resolve() if args.output else None,
            template_path_override=Path(args.template).resolve() if args.template else None,
            mapping_path=Path(args.mapping).resolve() if args.mapping else None,
            process_output_override=Path(args.process_output).resolve() if args.process_output else None,
        )
    except FpaPayloadError as exc:
        print(f"FPA Excel 生成失败: {exc}")
        return 2
    print(json.dumps({
        "output_path": process["output_path"],
        "process_output_path": process["process_output_path"],
        "item_count": process["item_count"],
        "adjusted_fp_total": process["estimates"]["adjusted_fp_total"],
        "work_days_middle": process["estimates"]["work_days"]["middle"],
        "quality_gate": process["quality_gate"]["status"],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
