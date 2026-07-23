from __future__ import annotations

import argparse
import re
import secrets
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - depends on runtime environment
    print("缺少 openpyxl 依赖，请先安装后端依赖或在 Docker 容器内执行。", file=sys.stderr)
    raise SystemExit(2) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = PROJECT_ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.config import get_config
from app.db import hash_password, initialize_database, open_connection, utc_now


ROLE_ALIASES = {
    "admin": "admin",
    "管理员": "admin",
    "user": "user",
    "普通用户": "user",
}
USERNAME_RE = re.compile(r"^[A-Za-z0-9_.-]+$")


@dataclass(frozen=True)
class UserRow:
    row_number: int
    display_name: str
    username: str
    password: str
    role: str
    default_system_code: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从 Excel 批量导入 TeamTools 用户")
    parser.add_argument("excel_path", help="用户初始化 Excel 路径")
    parser.add_argument("--sheet", help="工作表名称；不传则读取第一个工作表")
    parser.add_argument("--has-header", action="store_true", help="Excel 第一行为表头时启用")
    parser.add_argument("--admin-username", action="append", default=[], help="指定管理员用户名，可重复传入")
    parser.add_argument("--admin-display-name", action="append", default=[], help="指定管理员姓名，可重复传入")
    parser.add_argument("--default-system-code", help="Excel 未提供默认系统时使用的默认系统编码")
    parser.add_argument("--update", action="store_true", help="账号已存在时更新密码、角色、展示名和默认系统")
    parser.add_argument("--dry-run", action="store_true", help="只预览将导入的账号，不写数据库")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_role(raw: str, *, display_name: str, username: str, admin_names: set[str], admin_users: set[str]) -> str:
    explicit = ROLE_ALIASES.get(raw.strip().lower()) or ROLE_ALIASES.get(raw.strip())
    if explicit:
        return explicit
    if username.lower() in admin_users or display_name in admin_names:
        return "admin"
    return "user"


def load_rows(args: argparse.Namespace) -> list[UserRow]:
    path = Path(args.excel_path).expanduser()
    if not path.exists():
        raise ValueError(f"Excel 文件不存在：{path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            raise ValueError(f"工作表不存在：{args.sheet}")
        sheet = workbook[args.sheet]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    start_row = 2 if args.has_header else 1
    admin_users = {item.strip().lower() for item in args.admin_username if item.strip()}
    admin_names = {item.strip() for item in args.admin_display_name if item.strip()}
    rows: list[UserRow] = []
    seen: set[str] = set()
    for row_number in range(start_row, sheet.max_row + 1):
        values = [cell_text(sheet.cell(row=row_number, column=column).value) for column in range(1, 6)]
        if not any(values):
            continue

        display_name, username, password = values[:3]
        raw_role = values[3]
        default_system_code = values[4] or args.default_system_code

        if not display_name:
            raise ValueError(f"第 {row_number} 行缺少姓名")
        if not username:
            raise ValueError(f"第 {row_number} 行缺少用户名")
        if not USERNAME_RE.fullmatch(username):
            raise ValueError(f"第 {row_number} 行用户名只能包含字母、数字、下划线、点和短横线")
        if username.lower() in seen:
            raise ValueError(f"Excel 中用户名重复：{username}")
        if not password:
            raise ValueError(f"第 {row_number} 行缺少初始化密码")
        if len(password) < 8:
            raise ValueError(f"第 {row_number} 行初始化密码少于 8 位")

        seen.add(username.lower())
        rows.append(
            UserRow(
                row_number=row_number,
                display_name=display_name,
                username=username,
                password=password,
                role=normalize_role(
                    raw_role,
                    display_name=display_name,
                    username=username,
                    admin_names=admin_names,
                    admin_users=admin_users,
                ),
                default_system_code=default_system_code or None,
            )
        )
    return rows


def print_preview(rows: list[UserRow]) -> None:
    print(f"待导入账号数：{len(rows)}")
    print("行号\t用户名\t姓名\t角色\t默认系统\t密码")
    for row in rows:
        default_system = row.default_system_code or "-"
        print(f"{row.row_number}\t{row.username}\t{row.display_name}\t{row.role}\t{default_system}\t已提供")


def import_rows(rows: list[UserRow], *, update_existing: bool) -> tuple[int, int, int]:
    config = get_config()
    initialize_database(config.db_path, seed_dev_users_enabled=False)
    created = 0
    updated = 0
    skipped = 0
    with open_connection(config.db_path) as conn:
        for row in rows:
            existing = conn.execute("SELECT id FROM users WHERE username = ?", (row.username,)).fetchone()
            now = utc_now()
            if existing and not update_existing:
                skipped += 1
                continue
            if existing:
                conn.execute(
                    """
                    UPDATE users
                    SET display_name = ?, password_hash = ?, role = ?, default_system_code = ?,
                        initial_password_seed = ?,
                        enabled = 1, updated_at = ?
                    WHERE username = ?
                    """,
                    (
                        row.display_name,
                        hash_password(row.password),
                        row.role,
                        row.default_system_code,
                        row.password,
                        now,
                        row.username,
                    ),
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO users(id, username, display_name, password_hash, role,
                                      default_system_code, initial_password_seed, enabled, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        "user-" + secrets.token_hex(8),
                        row.username,
                        row.display_name,
                        hash_password(row.password),
                        row.role,
                        row.default_system_code,
                        row.password,
                        now,
                        now,
                    ),
                )
                created += 1
        conn.commit()
    return created, updated, skipped


def main() -> int:
    args = parse_args()
    try:
        rows = load_rows(args)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    print_preview(rows)
    admins = [row.username for row in rows if row.role == "admin"]
    if not admins:
        print("未识别到管理员账号，请使用 --admin-username 或 --admin-display-name 指定。", file=sys.stderr)
        return 2

    if args.dry_run:
        print("dry-run 模式未写入数据库。")
        return 0

    created, updated, skipped = import_rows(rows, update_existing=args.update)
    print(f"导入完成：新增 {created}，更新 {updated}，跳过 {skipped}。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
