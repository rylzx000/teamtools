import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
SCRIPT_PATH = ROOT_DIR / "scripts" / "fpa" / "fill_fpa_workbook.py"
SAMPLE_PAYLOAD_PATH = (
    ROOT_DIR
    / "data"
    / "modules"
    / "fpa"
    / "examples"
    / "excel"
    / "Excel脚本输入payload.sample.json"
)
TEMPLATE_PATH = ROOT_DIR / "data" / "modules" / "fpa" / "profile" / "templates" / "fpa_template.xlsx"

REQUIRED_SHEETS = ["项目特征", "规模估算", "开发费用估算", "模板使用说明&基础参数"]


class FillFpaWorkbookTest(unittest.TestCase):
    def run_script(self, payload_path, output_path, process_path):
        result = subprocess.run(
            [
                sys.executable,
                str(SCRIPT_PATH),
                "--payload",
                str(payload_path),
                "--output",
                str(output_path),
                "--process-output",
                str(process_path),
            ],
            cwd=ROOT_DIR,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            check=False,
        )
        self.assertEqual(
            result.returncode,
            0,
            msg=f"stdout:\n{result.stdout}\nstderr:\n{result.stderr}",
        )

    def assert_basic_workbook_contract(self, output_path, process_path, payload_path):
        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 10_000)
        self.assertTrue(process_path.exists())

        with zipfile.ZipFile(output_path) as workbook_zip:
            names = set(workbook_zip.namelist())
        self.assertIn("xl/styles.xml", names)
        self.assertIn("xl/workbook.xml", names)

        workbook = load_workbook(output_path, data_only=False)
        self.assertEqual(workbook.sheetnames, REQUIRED_SHEETS)

        formula_count = 0
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        self.assertGreater(formula_count, 0)

        payload = json.loads(payload_path.read_text(encoding="utf-8"))
        process = json.loads(process_path.read_text(encoding="utf-8"))
        self.assertEqual(process["schema_version"], "fpa.excel_process.v1")
        self.assertEqual(process["item_count"], len(payload["items"]))
        self.assertIn("middle", process["estimates"]["work_days"])
        self.assertIsInstance(process["estimates"]["work_days"]["middle"], (int, float))
        self.assertIn("status", process["quality_gate"])
        self.assertIn("deliverable_valid", process)

        return workbook, payload, process

    def make_payload_with_item_count(self, item_count):
        source_payload = json.loads(SAMPLE_PAYLOAD_PATH.read_text(encoding="utf-8"))
        payload = deepcopy(source_payload)
        source_items = source_payload["items"]
        items = []
        for index in range(item_count):
            item = deepcopy(source_items[index % len(source_items)])
            item["stable_id"] = f"FP-{index + 1:03d}"
            item["function_description"] = f"{item['function_description']}（扩展样例{index + 1}）"
            item["count_item_name"] = f"{item['count_item_name']}{index + 1:02d}"
            item["remark"] = f"{item['remark']}；扩展样例{index + 1}"
            items.append(item)
        payload["items"] = items
        return payload

    def test_generates_template_workbook_and_process_json(self):
        with tempfile.TemporaryDirectory() as temp_dir_name:
            temp_dir = Path(temp_dir_name)
            output_path = temp_dir / "fpa-review.xlsx"
            process_path = temp_dir / "fpa-review-process.json"

            self.run_script(SAMPLE_PAYLOAD_PATH, output_path, process_path)
            self.assert_basic_workbook_contract(output_path, process_path, SAMPLE_PAYLOAD_PATH)

    def test_project_features_preserve_template_defaults_except_c1_and_c7(self):
        with tempfile.TemporaryDirectory() as temp_dir_name:
            temp_dir = Path(temp_dir_name)
            output_path = temp_dir / "fpa-review.xlsx"
            process_path = temp_dir / "fpa-review-process.json"

            self.run_script(SAMPLE_PAYLOAD_PATH, output_path, process_path)

            template = load_workbook(TEMPLATE_PATH, data_only=False)
            generated = load_workbook(output_path, data_only=False)
            template_project = template["项目特征"]
            generated_project = generated["项目特征"]

            payload = json.loads(SAMPLE_PAYLOAD_PATH.read_text(encoding="utf-8"))
            expected_c1 = payload["project_features"]["规模计数时机"]
            expected_c7 = payload["project_features"]["完整性级别"]
            self.assertEqual(generated_project["C1"].value, expected_c1)
            self.assertEqual(generated_project["C7"].value, expected_c7)
            for row in range(2, 11):
                if row == 7:
                    continue
                self.assertEqual(
                    generated_project[f"C{row}"].value,
                    template_project[f"C{row}"].value,
                    msg=f"项目特征!C{row} should preserve template value",
                )
            for forbidden_value in ["中等", "3GL"]:
                self.assertNotIn(
                    forbidden_value,
                    [generated_project[f"C{row}"].value for row in range(3, 10) if row != 7],
                )

    def test_eleven_items_dynamic_rows_and_structured_remarks(self):
        with tempfile.TemporaryDirectory() as temp_dir_name:
            temp_dir = Path(temp_dir_name)
            source_payload = json.loads(SAMPLE_PAYLOAD_PATH.read_text(encoding="utf-8"))
            payload = deepcopy(source_payload)
            payload["items"] = payload["items"][:11]
            payload_path = temp_dir / "payload-11.json"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            output_path = temp_dir / "fpa-review-11.xlsx"
            process_path = temp_dir / "fpa-review-11-process.json"

            self.run_script(payload_path, output_path, process_path)
            workbook, checked_payload, process = self.assert_basic_workbook_contract(
                output_path,
                process_path,
                payload_path,
            )
            self.assertEqual(process["item_count"], 11)
            self.assertEqual(process["summary_row"], 17)
            self.assertEqual(process["detail_range"], "规模估算!B6:N16")

            template = load_workbook(TEMPLATE_PATH, data_only=False)
            size = workbook["规模估算"]
            cost = workbook["开发费用估算"]
            template_size = template["规模估算"]
            self.assertEqual(size.max_column, template_size.max_column)

            for offset, item in enumerate(checked_payload["items"]):
                row = 6 + offset
                expected_values = {
                    "B": item["system"],
                    "C": item["level1_module"],
                    "D": item["level2_module"],
                    "E": item["level3_module"],
                    "F": item["level4_module"],
                    "G": item["function_description"],
                    "H": item["count_item_name"],
                    "I": item["category"],
                    "K": item["reuse"],
                    "L": item["change_type"],
                }
                for column, expected in expected_values.items():
                    actual = size[f"{column}{row}"].value
                    if expected == "":
                        self.assertIn(actual, ("", None), msg=f"{column}{row}")
                    else:
                        self.assertEqual(actual, expected, msg=f"{column}{row}")

                remark = size[f"N{row}"].value
                self.assertIsInstance(remark, str)
                for label in ["类别原因：", "复用原因：", "修改类型原因："]:
                    self.assertIn(label, remark)
                self.assertIn(item["category"], remark)
                self.assertIn(item["reuse"], remark)
                self.assertIn(item["change_type"], remark)

                self.assertIsInstance(size[f"A{row}"].value, str)
                self.assertTrue(size[f"A{row}"].value.startswith("="))
                self.assertIsInstance(size[f"J{row}"].value, str)
                self.assertTrue(size[f"J{row}"].value.startswith("="))
                self.assertIsInstance(size[f"M{row}"].value, str)
                self.assertTrue(size[f"M{row}"].value.startswith("="))

            self.assertEqual(size["A17"].value, "合计")
            self.assertEqual(size["C2"].value, "=J17")
            self.assertEqual(size["C3"].value, "=M17")
            self.assertEqual(size["J17"].value, "=SUM(J6:J16)")
            self.assertEqual(size["M17"].value, "=SUM(M6:M16)")
            self.assertEqual(cost["C1"].value, "='规模估算'!M17")

            validations = [str(dv.sqref) + ":" + str(dv.formula1) for dv in size.data_validations.dataValidation]
            self.assertTrue(any('I6:I16:"ILF,EIF,EI,EO,EQ"' in validation for validation in validations))
            self.assertTrue(any('K6:K16:"高,中,低"' in validation for validation in validations))
            self.assertTrue(any('L6:L16:"新增,修改,删除"' in validation for validation in validations))
            self.assertFalse(any("I6:I17" in validation for validation in validations))
            self.assertFalse(any("K6:K17" in validation for validation in validations))
            self.assertFalse(any("L6:L17" in validation for validation in validations))

            merged_ranges = {str(merged) for merged in size.merged_cells.ranges}
            self.assertIn("A17:I17", merged_ranges)
            self.assertIn("K17:L17", merged_ranges)
            self.assertTrue({"A2:B2", "A4:N4", "D1:N1", "D2:N2", "A1:B1", "A3:B3", "D3:N3"}.issubset(merged_ranges))

    def test_thirty_and_fifty_items_expand_in_single_workbook(self):
        for item_count in (30, 50):
            with self.subTest(item_count=item_count):
                with tempfile.TemporaryDirectory() as temp_dir_name:
                    temp_dir = Path(temp_dir_name)
                    payload = self.make_payload_with_item_count(item_count)
                    payload_path = temp_dir / f"payload-{item_count}.json"
                    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
                    output_path = temp_dir / f"fpa-review-{item_count}.xlsx"
                    process_path = temp_dir / f"fpa-review-{item_count}-process.json"

                    self.run_script(payload_path, output_path, process_path)
                    workbook, checked_payload, process = self.assert_basic_workbook_contract(
                        output_path,
                        process_path,
                        payload_path,
                    )

                    last_detail_row = 5 + item_count
                    summary_row = 6 + item_count
                    self.assertEqual(process["item_count"], item_count)
                    self.assertEqual(process["summary_row"], summary_row)
                    self.assertEqual(process["detail_range"], f"规模估算!B6:N{last_detail_row}")

                    size = workbook["规模估算"]
                    cost = workbook["开发费用估算"]
                    last_item = checked_payload["items"][-1]
                    self.assertEqual(size[f"B{last_detail_row}"].value, last_item["system"])
                    self.assertEqual(size[f"C{last_detail_row}"].value, last_item["level1_module"])
                    self.assertEqual(size[f"D{last_detail_row}"].value, last_item["level2_module"])
                    self.assertEqual(size[f"H{last_detail_row}"].value, last_item["count_item_name"])
                    self.assertEqual(size[f"I{last_detail_row}"].value, last_item["category"])
                    self.assertEqual(size[f"K{last_detail_row}"].value, last_item["reuse"])
                    self.assertEqual(size[f"L{last_detail_row}"].value, last_item["change_type"])
                    self.assertEqual(size[f"A{summary_row}"].value, "合计")
                    self.assertEqual(size["C2"].value, f"=J{summary_row}")
                    self.assertEqual(size["C3"].value, f"=M{summary_row}")
                    self.assertEqual(size[f"J{summary_row}"].value, f"=SUM(J6:J{last_detail_row})")
                    self.assertEqual(size[f"M{summary_row}"].value, f"=SUM(M6:M{last_detail_row})")
                    self.assertEqual(cost["C1"].value, f"='规模估算'!M{summary_row}")

                    for column in ("A", "J", "M"):
                        self.assertIsInstance(size[f"{column}{last_detail_row}"].value, str)
                        self.assertTrue(size[f"{column}{last_detail_row}"].value.startswith("="))
                    for column in ("B", "C", "D", "I", "K", "L"):
                        self.assertEqual(
                            size[f"{column}{last_detail_row}"].style_id,
                            size[f"{column}6"].style_id,
                            msg=f"{column}{last_detail_row} should keep detail-row style",
                        )

                    validations = [str(dv.sqref) + ":" + str(dv.formula1) for dv in size.data_validations.dataValidation]
                    self.assertTrue(
                        any(
                            f'I6:I{last_detail_row}:"ILF,EIF,EI,EO,EQ"' in validation
                            for validation in validations
                        )
                    )
                    self.assertTrue(any(f'K6:K{last_detail_row}:"高,中,低"' in validation for validation in validations))
                    self.assertTrue(
                        any(f'L6:L{last_detail_row}:"新增,修改,删除"' in validation for validation in validations)
                    )
                    self.assertFalse(any(f"I6:I{summary_row}" in validation for validation in validations))
                    self.assertFalse(any(f"K6:K{summary_row}" in validation for validation in validations))
                    self.assertFalse(any(f"L6:L{summary_row}" in validation for validation in validations))

                    merged_ranges = {str(merged) for merged in size.merged_cells.ranges}
                    self.assertIn(f"A{summary_row}:I{summary_row}", merged_ranges)
                    self.assertIn(f"K{summary_row}:L{summary_row}", merged_ranges)


if __name__ == "__main__":
    unittest.main()
